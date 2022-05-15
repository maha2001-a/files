from hashlib import sha256
import base64
from telnetlib import NOP
from Crypto import Random
from Crypto.Cipher import AES
import hashlib
import random
from Crypto.Util.Padding import pad, unpad
from openpyxl import load_workbook
from base64 import b64encode, b64decode
from hashlib import md5
from Crypto.Random import get_random_bytes

wb = load_workbook(f"Excel_Sheets/Encrypted.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['B']  # Column
encrypted_list = [column[x].value for x in range(len(column))]
column = ws['A']  # Column
hashed_list = [column[x].value for x in range(len(column))]

word=input("Enter word to find..")
BS = 32
padd = lambda s: bytes(s + (BS - len(s) % BS) * chr(BS - len(s) % BS), 'utf-8')
unpadd = lambda s : s[0:-ord(s[-1:])]

class AESCipher1:

    def __init__( self, key ):
        self.key = sha256(key.encode('utf8')).digest()


    def encryptt( self, raw ):
        raw = padd(raw)
        iv = Random.new().read( AES.block_size )
        cipher = AES.new(self.key, AES.MODE_CBC, iv )
        return base64.b64encode( iv + cipher.encrypt( raw ) )

    def decryptt( self, enc ):
        enc = base64.b64decode(enc)
        iv = enc[:16]
        cipher = AES.new(self.key, AES.MODE_CBC, iv )
        return unpadd(cipher.decrypt( enc[16:] )).decode('utf8')
seedd=input("enter password..")
pwd=random.seed(seedd)
random.random=random.getrandbits(256)
pwd=random.getrandbits(256)
pwd=str(pwd)
#print(len(pwd))
pwd = hashlib.sha256(pwd.encode())
pwd=pwd.hexdigest()
#print("Password length..",len(pwd))
random.random=random.getrandbits(256)
session_key=random.getrandbits(256)
session_key=str(session_key)
cipher = AESCipher1(session_key)
encrypted_text = cipher.encryptt(word)
encrypted_text = encrypted_text.decode('utf-8')
#decrypted_text = cipher.decrypt(encrypted_text)
print(len(encrypted_text),"Encrypted Text....",encrypted_text)
#print(decrypted_text)

#Hashing.......
findword_hashed = hashlib.sha256(word.encode('utf-8')).hexdigest()
findword_hashed=str(findword_hashed)
print(len(findword_hashed),"Hash..",findword_hashed)
#Coverting to binary
binary_converted_a = ''.join(format(ord(c), 'b') for c in findword_hashed)
binary_converted_b = ''.join(format(ord(c), 'b') for c in encrypted_text)
if (len(binary_converted_a)>=len(binary_converted_b)):
    n=len(binary_converted_a)
    binary_converted_b=binary_converted_b.ljust(n,'0')
else:
    l=len(binary_converted_b)
    binary_converted_a=binary_converted_a.ljust(l,'0')
#binary_converted_a=binary_converted_a.ljust(60, '0')
#binary_converted_b=binary_converted_b.ljust(60,'0')
print(binary_converted_a)
print(binary_converted_b)
def xor_two_str(a,b):
    xored = []
    for i in range(max(len(a), len(b))):
        xored_value = ord(a[i%len(a)]) ^ ord(b[i%len(b)])
        xored.append(hex(xored_value)[2:])
    return ''.join(xored)
    
c=xor_two_str(binary_converted_a,binary_converted_b)
print(c)
o=xor_two_str(binary_converted_b,c)
print(o==binary_converted_a)

d = hashlib.sha256(binary_converted_b.encode('utf-8')).hexdigest()
d=str(d)
for g in hashed_list:
    a=g
    a = ''.join(format(ord(c), 'b') for c in a)
    if (len(a)>=len(c)):
       n=len(a)
       c=c.ljust(n,'0')
    else:
       l=len(c)
       a=a.ljust(l,'0')
    xoring=xor_two_str(a,c)
    hashing=hashlib.sha256(xoring.encode('utf-8')).hexdigest()
    hashing=str(hashing)
    counter=0
    if(d==hashing):
        print("found at index ",g,hashed_list.index(g))
        found_index=hashed_list.index(g)
    else:
        NOP
class AESCipher:
   def __init__(self, key):
        self.key = md5(key.encode('utf8')).digest()

   def encrypt(self, data):
        iv = get_random_bytes(AES.block_size)
        self.cipher = AES.new(self.key, AES.MODE_CBC, iv)
        return b64encode(iv + self.cipher.encrypt(pad(data.encode('utf-8'), AES.block_size)))

   def decrypt(self, data):
        raw = b64decode(data)
        self.cipher = AES.new(self.key, AES.MODE_CBC, raw[:AES.block_size])
        return unpad(self.cipher.decrypt(raw[AES.block_size:]), AES.block_size)


if __name__ == '__main__':
    found_file=encrypted_list[found_index]
    found_file=found_file.encode('utf-8')
    decypted_file=AESCipher(pwd).decrypt(found_file).decode('utf-8')
    
   # encrypted_file=AESCipher(pwd).decrypt(decypted_file).decode('utf-8')
    print(decypted_file)
        