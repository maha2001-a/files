from fileinput import filename
from cryptography.fernet import Fernet
import os
key=Fernet.generate_key()
print(key)
fernet=Fernet(key)
with open('key.key','wb') as filekey:
    filekey.write(key)
with open('key.key','rb') as filekey:
    key=filekey.read()
#filename=input("enter filename.")
dir="E:/fyp2/codes/index_table_generation/files/AudioFiles"
dir_list = os.listdir(dir)
for y in dir_list:
    path=y
    with open(f'AudioFiles/{y}','rb') as file:
      originalaudio=file.read()
      encrypted=fernet.encrypt(originalaudio)
    with open(f"Encrypted_Audio/{path}",'wb') as encrypted_file:
       encrypted_file.write(encrypted)
       fernet=Fernet(key)
dir="E:/fyp2/codes/index_table_generation/files/Encrypted_Audio"
dir_list = os.listdir(dir)
for k in dir_list:
    with open(f"Encrypted_Audio/{k}",'rb') as enc_file:
       encrypted=enc_file.read()
       decrypted=fernet.decrypt(encrypted)
    with open(f"Decrypted_Audio/{k}",'wb') as dec_file:
        dec_file.write(decrypted)
# importing libraries 
import speech_recognition as sr 
import os 
from pydub import AudioSegment
from pydub.silence import split_on_silence

# create a speech recognition object
r = sr.Recognizer()
path_dir="E://fyp2//codes//index_table_generation//files//AudioFiles"
dir_list = os.listdir(path_dir)
#print(dir_list)
print("Let Start Converting Audio to Text.....")
for x in dir_list:
# a function that splits the audio file into chunks
# and applies speech recognition
    path=x
    def get_large_audio_transcription(path):
   # """
   # Splitting the large audio file into chunks
   # and apply speech recognition on each of these chunks
   # """
    # open the audio file using pydub
       sound = AudioSegment.from_wav(f"AudioFiles/{path}")  
    # split audio sound where silence is 700 miliseconds or more and get chunks
       chunks = split_on_silence(sound,
        # experiment with this value for your target audio file
            min_silence_len = 500,
        # adjust this per requirement
            silence_thresh = sound.dBFS-14,
        # keep the silence for 1 second, adjustable as well
            keep_silence=500,
    )
       folder_name = "zaudio-chunks"
    # create a directory to store the audio chunks
       if not os.path.isdir(folder_name):
          os.mkdir(folder_name)
       whole_text = ""
    # process each chunk 
       for i, audio_chunk in enumerate(chunks, start=1):
        # export audio chunk and save it in
        # the `folder_name` directory.
           chunk_filename = os.path.join(folder_name, f"chunk{i}.wav")
           audio_chunk.export(chunk_filename, format="wav")
        # recognize the chunk
           with sr.AudioFile(chunk_filename) as source:
                audio_listened = r.record(source)
            # try converting it to text
                try:
                    text = r.recognize_google(audio_listened)
                except sr.UnknownValueError as e:
                    print("Error:", str(e))
                else:
                    text = f"{text.capitalize()}. "
                   # print(chunk_filename, ":", text)
                    whole_text += text
    # return the text for all chunks detected
       return whole_text


 
#print("Files and directories in '", path_dir, "' :")
 
# print all files

    path = x
    with open(f"Text_Files/{path}.txt","w")as external_file:
        add_text= get_large_audio_transcription(path)
        print(add_text,file=external_file)
        external_file.close()

    get_large_audio_transcription(path)
   # print("Just Hold On .... ")
print("DONE...")


from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.stem import LancasterStemmer
import os
print("Let do stemming..")
path_dir="E:/fyp2/codes/index_table_generation/files/Text_Files"
dir_list = os.listdir(path_dir)
for x in dir_list:
    text = open(f"Text_Files/{x}").read().lower()
    stopWords = set(stopwords.words('english'))
    stopWords.update((',','.',':','and','I','A','And','So','arnt','This','When','It','many','Many','so','cant','Yes','yes','No','no','These','these','his'))
    stopWords.update(('his','one','could','try','may'))
#stopWords = set(nltk.corpus.stopwords.words('english'))
#stopWords.append('.')
    words = word_tokenize(text)
    wordsFiltered = []

    for w in words:
       if w not in stopWords:
          wordsFiltered.append(w)
    def listToString(s): 
    
    # initialize an empty string
        str1 = " " 
    
    # return string  
        return (str1.join(s))

    string=listToString(wordsFiltered)
#print(wordsFiltered)
    f = open(f"Stemming/{x}", "w")
    f.write(string)
    f.close()
print("Finally DONE...")
from encodings import utf_8
from importlib.metadata import files
from msilib.schema import File
import os
from collections import Counter
from telnetlib import NOP
from xlsxwriter import Workbook
from hashlib import md5
from base64 import b64decode
import random
from hashlib import sha256
import base64
from Crypto import Random
from Crypto.Cipher import AES
from base64 import b64encode
import hashlib
import hmac
from Crypto.Random import get_random_bytes
from Crypto.Util.Padding import pad, unpad
print("Index Table is being generated here..")
path_dir="E:/fyp2/codes/index_table_generation/files/Stemming"
dir_list = os.listdir(path_dir)
path_dir_audiofiles="E:/fyp2/codes/index_table_generation/files/AudioFiles"
dir_list_audiofiles = os.listdir(path_dir_audiofiles)
max_word=[]
file_list=[]
audio_file_list=[]
def most_frequent(List):
    counter = 0
    num = List[0]
    for i in List:
        curr_frequency = List.count(i)
        if(curr_frequency> counter):
            counter = curr_frequency
            num = i
    return num
for x in dir_list_audiofiles:
    #define text file to open
    my_file_audio = open(f"AudioFiles/{x}", 'r')
    audio_file_list.append(x)
for x in dir_list:
    #define text file to open
    my_file = open(f"Stemming/{x}", 'r')
    file_list.append(x)
    for line in my_file:
        for i in dir_list:
           i=line.split()
        #print(i)
        max_count=most_frequent(i)
        max_word.append(max_count)
    #print(max_word)
    #print(file_list)

workbook = Workbook(f"Excel_Sheets/Ecl.xlsx")
Report_Sheet = workbook.add_worksheet()
# Write the column data.
Report_Sheet.write_column(0, 0, max_word)
Report_Sheet.write_column(0, 1, audio_file_list)

workbook.close()
# Write the column headers if required.
#Report_Sheet.write(0, 0, 'Key Words')
#Report_Sheet.write(0, 1, 'File Names')
seedd=input("enter password..")
f = open("key.txt", "w")
f.write(seedd)
f.close()
pwd=random.seed(seedd)
pwd=random.getrandbits(256)
pwd=str(pwd)
#print(pwd)
#print(len(pwd))
pwd = hashlib.sha256(pwd.encode())
pwd=pwd.hexdigest()
pwd=str(pwd)
#print("Password length..",len(pwd))
#read text file into list
   # data = my_file.read()
   # data_into_list= data.replace('\n', ' ').split(".")
  
# printing the data
   # print(data_into_list)
class AESCipher:
   def __init__(self, key):
        self.key = md5(key.encode('utf8')).digest()

   def encrypt(self, data):
        iv = get_random_bytes(AES.block_size)
        self.cipher = AES.new(self.key, AES.MODE_CBC, iv)
        return b64encode(iv + self.cipher.encrypt(pad(data.encode('utf-8'), AES.block_size)))

   def decrypt(self, data):
        raw = b64decode(data)
        self.cipher = AES.new(self.key, AES.MODE_CBC, raw[:AES.block_size])
        return unpad(self.cipher.decrypt(raw[AES.block_size:]), AES.block_size)


if __name__ == '__main__':
    pwd=str(pwd)
    Encrypted_list=[]
    Decrypted_list=[]
    #print(pwd)
    for i in audio_file_list:
        encrypted_text= AESCipher(pwd).encrypt(i).decode('utf-8')
        #print(encrypted_text)
        Encrypted_list.append(encrypted_text)
    #print("Encrypted_list",Encrypted_list)
    for j in Encrypted_list:
        decypted_text=AESCipher(pwd).decrypt(j).decode('utf-8')
        Decrypted_list.append(decypted_text)
    #print("Decrypted list",Decrypted_list)
hashed_list=[]
pwd=pwd.encode('utf-8')
for i in max_word :
    i=i.encode('utf-8')
    result = hmac.new(i, pwd, hashlib.sha256)
    reslt=result.hexdigest()
    hashed_list.append(reslt)
#print(hashed_list)
workbook2 = Workbook(f"Excel_Sheets/Encrypted.xlsx")
Report_Sheet = workbook2.add_worksheet()

# Write the column headers if required.
#Report_Sheet.write(0, 0, 'Key Words')
#Report_Sheet.write(0, 1, 'File Names')


# Write the column data.
Report_Sheet.write_column(0, 0, hashed_list)
Report_Sheet.write_column(0, 1, Encrypted_list)
my_file.close()
workbook2.close()
print("Table Generated..")

from hashlib import sha256
import base64
from telnetlib import NOP
from Crypto import Random
from Crypto.Cipher import AES
import hashlib
import sys
import hmac
import random
from Crypto.Util.Padding import pad, unpad
from openpyxl import load_workbook
from base64 import b64encode, b64decode
from hashlib import md5
from fileinput import filename
from cryptography.fernet import Fernet
from Crypto.Random import get_random_bytes
wb = load_workbook(f"Excel_Sheets/Encrypted.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['B']  # Column
encrypted_list = [column[x].value for x in range(len(column))]
column = ws['A']  # Column
hashed_list = [column[x].value for x in range(len(column))]
#searching windows(enter keyword to search)
word=input("Enter word to find..")
BS = 32
padd = lambda s: bytes(s + (BS - len(s) % BS) * chr(BS - len(s) % BS), 'utf-8')
unpadd = lambda s : s[0:-ord(s[-1:])]

class AESCipher1:

    def __init__( self, key ):
        self.key = sha256(key.encode('utf8')).digest()


    def encryptt( self, raw ):
        raw = padd(raw)
        iv = Random.new().read( AES.block_size )
        cipher = AES.new(self.key, AES.MODE_CBC, iv )
        return base64.b64encode( iv + cipher.encrypt( raw ) )

    def decryptt( self, enc ):
        enc = base64.b64decode(enc)
        iv = enc[:16]
        cipher = AES.new(self.key, AES.MODE_CBC, iv )
        return unpadd(cipher.decrypt( enc[16:] )).decode('utf8')
import secrets
sys_random = secrets.SystemRandom()
f = open("key.txt", "r")
key=f.read()
#Decryption Password
seedd=input("enter password..")
if(seedd!=key):
    print("Wrong Password..")
    sys.exit()
pwd=random.seed(seedd)
pwd=random.getrandbits(256)
pwd=str(pwd)
#print(pwd)
#print(len(pwd))
pwd = hashlib.sha256(pwd.encode())
pwd=pwd.hexdigest()
pwd=str(pwd)
#print("Password length..",len(pwd))
session_key=sys_random.getrandbits(256)
session_key=str(session_key)
#print(session_key)
cipher = AESCipher1(session_key)
encrypted_text = cipher.encryptt(word)
encrypted_text = encrypted_text.decode('utf-8')
#decrypted_text = cipher.decrypt(encrypted_text)
#print(len(encrypted_text),"Encrypted Text....",encrypted_text)
#print(decrypted_text)

#Hashing.......
word=word.encode('utf-8')
pwd=pwd.encode('utf-8')
findword_hashed = hmac.new(word, pwd, hashlib.sha256)
findword_hashed=findword_hashed.hexdigest()
findword_hashed=str(findword_hashed)
#print("Hash..",findword_hashed)
#Coverting to binary
binary_converted_a = ''.join(format(ord(c), 'b') for c in findword_hashed)
binary_converted_b = ''.join(format(ord(c), 'b') for c in encrypted_text)
if (len(binary_converted_a)>=len(binary_converted_b)):
    n=len(binary_converted_a)
    binary_converted_b=binary_converted_b.ljust(n,'0')
else:
    l=len(binary_converted_b)
    binary_converted_a=binary_converted_a.ljust(l,'0')
#binary_converted_a=binary_converted_a.ljust(60, '0')
#binary_converted_b=binary_converted_b.ljust(60,'0')
#print(binary_converted_a)
#print(binary_converted_b)
def xor_two_str(a,b):
    xored = []
    for i in range(max(len(a), len(b))):
        xored_value = ord(a[i%len(a)]) ^ ord(b[i%len(b)])
        xored.append(hex(xored_value)[2:])
    return ''.join(xored)
    
c=xor_two_str(binary_converted_a,binary_converted_b)
#print(c)
o=xor_two_str(binary_converted_b,c)
#print(o==binary_converted_a)

d = hashlib.sha256(binary_converted_b.encode('utf-8')).hexdigest()
d=str(d)
for g in hashed_list:
    a=g
    a = ''.join(format(ord(c), 'b') for c in a)
    if (len(a)>=len(c)):
       n=len(a)
       c=c.ljust(n,'0')
    else:
       l=len(c)
       a=a.ljust(l,'0')
    xoring=xor_two_str(a,c)
    hashing=hashlib.sha256(xoring.encode('utf-8')).hexdigest()
    hashing=str(hashing)
    counter=0
    if(d==hashing):
        #print("found at index ",g,hashed_list.index(g))
        found_index=hashed_list.index(g)
        class AESCipher:
            def __init__(self, key):
                self.key = md5(key.encode('utf8')).digest()

            def encrypt(self, data):
                iv = get_random_bytes(AES.block_size)
                self.cipher = AES.new(self.key, AES.MODE_CBC, iv)
                return b64encode(iv + self.cipher.encrypt(pad(data.encode('utf-8'), AES.block_size)))

            def decrypt(self, data):
                raw = b64decode(data)
                self.cipher = AES.new(self.key, AES.MODE_CBC, raw[:AES.block_size])
                return unpad(self.cipher.decrypt(raw[AES.block_size:]), AES.block_size)
        bool=0


        if __name__ == '__main__':
            pwd=pwd.decode('utf-8')
            #print(pwd)
            found_file=encrypted_list[found_index]
            found_file=found_file.encode('utf-8')
            decypted_file=AESCipher(pwd).decrypt(found_file).decode('utf-8')
    
   # encrypted_file=AESCipher(pwd).decrypt(decypted_file).decode('utf-8')
            print(decypted_file)
#print(decypted_file)

        with open('key.key','rb') as filekey:
            key=filekey.read()
            fernet=Fernet(key)
        with open(f"Encrypted_Audio/{decypted_file}",'rb') as enc_file:
            encrypted=enc_file.read()
            decrypted=fernet.decrypt(encrypted)
        with open(f"Result/{decypted_file}",'wb') as dec_file:
             dec_file.write(decrypted)
        break
    else:
        bool=1

if(bool==1):
    print("Keyword not found.")


